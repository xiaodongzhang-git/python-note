# # -*- coding: utf-8 -*-
# ModuleNotFoundError: No module named 'openpyxl' 第三方包要安装 pip install xxx
from openpyxl import Workbook, load_workbook

# 写入
wb = Workbook()
ws = wb.active
ws.title = "sheet0"
ws2 = wb.create_sheet("sheet2")
ws1 = wb.create_sheet("sheet1", 1) # 可以指定sheet插入位置
print(wb.sheetnames)
for ws in wb:
    for i in range(10):
        for j in range(10):
            ws.cell(row=i+1, column=j+1, value=ws.title + str(i+1) + str(j+1))

wb.save('test.xlsx')

# 读取
path = 'test.xlsx'
wb = load_workbook(path)
for ws in wb:
    for row in ws.rows:
        for cell in row:
            print(cell.value)
